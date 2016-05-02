from openpyxl import load_workbook
import os

match = raw_input("what match is this?")
teams_in = raw_input("what teams are in this match? (seperate by commas)")

wb = load_workbook(filename='C:\Users\Jake\Desktop\Scouting Scripts\Scouting.xlsx')

ws = wb['Match Scouting Data']

teams = teams_in.split(',',6)

for team in teams:
    match_data_file = open('C:\Users\Jake\Documents\Scouting\Team_' + team + "_Match_" + match + ".txt", "r")
    match_data = match_data_file.readlines()
    data = match_data[0].split('.', 22)

    #flip through all cells and find empty one
    for num in range(3,1001):
        if ws['C' + str(num)].value == None:
            print "row" + str(num) + "is empty"
            row = num
            break

    print row
    print data
    for numba in range (0,26):
        if numba == 1:
            #first name
            print "Scout Name: " + data[0]
            ws['AC' + str(row)] = data[0]
        if numba == 2:
            #Team num
            ws['C' + str(row)] = int(data[1])
        if numba == 3:
            #Match num
            ws['D' + str(row)] = int(data[2])
        if numba == 4:
            #auto start
            ws['E' + str(row)] = data[3]
        if numba == 5:
            #auto reach
            ws['F' + str(row)] = data[4]
            ws['H' + str(row)] = data[4]
        if numba == 6:
            #auto df crossed
            ws['G' + str(row)] = data[5]
        if numba == 7:
            #hi made auto
            ws['I' + str(row)] = int(data[6])
        if numba == 8:
            #lo made auto
            ws['I' + str(row)] = int(data[7])
        if numba == 9:
            #port c crosses
            ws['L' + str(row)] = int(data[8])
        if numba == 10:
            #lo bar crosses
            ws['K' + str(row)] = int(data[16])
        if numba == 11:
            #cheval crosses
            ws['M' + str(row)] = int(data[9])
        if numba == 12:
            #ramparts crosses
            ws['O' + str(row)] = int(data[10])
        if numba == 13:
            #moat crosses
            ws['N' + str(row)] = int(data[11])
        if numba == 14:
            #draw bridge crosses
            ws['P' + str(row)] = int(data[12])
        if numba == 15:
            #sally port crosses
            ws['Q' + str(row)] = int(data[13])
        if numba == 16:
            #rock wall crosses
            ws['R' + str(row)] = int(data[14])
        if numba == 17:
            #rough t crosses
            ws['S' + str(row)] = int(data[15])
        if numba == 18:
            #lo goals made
            ws['T' + str(row)] = int(data[19])
        if numba == 19:
            #hi fail
            ws['U' + str(row)] = int(data[18])
        if numba == 20:
            #hi made
            ws['V' + str(row)] = int(data[17])
        if numba == 21:
            #scale
            ws['X' + str(row)] = data[20]
        if numba == 22:
            #challenge
            ws['W' + str(row)] = data[21]
        if numba == 23:
            #comments
            ws['AB' + str(row)] = data[22]

wb.save('C:\Users\Jake\Desktop\Scouting Scripts\Scouting.xlsx')
